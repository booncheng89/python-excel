from openpyxl import Workbook

if __name__ == "__main__":
    wb = Workbook(write_only=True)

    # create state's worksheet and add random data
    state_ws = wb.create_sheet("Malaysia State")
    # state header
    state_ws.append(["id", "State Name"])
    # state data
    state_ws.append([1, "Sarawak"])
    state_ws.append([2, "Sabah"])
    state_ws.append([3, "Perlis"])
    state_ws.append([4, "Penang"])

    # create city's worksheet and add random data
    city_ws = wb.create_sheet("Malaysia City")
    # city header
    city_ws.append(["id", "City Name"])
    # city data
    city_ws.append([1, "Kuching"])
    city_ws.append([2, "Sibu"])
    city_ws.append([3, "Miri"])
    city_ws.append([4, "Kota Kinabalu"])

    #save file
    wb.save("sample2.xlsx")