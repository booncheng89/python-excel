from openpyxl import load_workbook

if __name__ == "__main__":
    wb = load_workbook(filename="sample1.xlsx", read_only=False)
    ws = wb["Sheet1"]
    # add Shah Alam
    ws["A7"] = 6
    ws["B7"] = "Shah Alam"

    # add Georgetown
    ws["A8"] = 7
    ws["B8"] = "Georgetown"
    wb.save("sample1.xlsx") # save to original file name
